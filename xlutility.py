import openpyxl


def function_operation():
    FILE = "C:\\Users\\Aashi\\PycharmProjects\\Parameterize_Example\\Multiply.xlsx"
    num_list = []
    wb = openpyxl.load_workbook(FILE)
    sh = wb["Sheet1"]
    row_count = sh.max_row
    # col_count = sh.max_column

    for r in range(2, row_count + 1):
        num1 = sh.cell(r, 1).value
        num2 = sh.cell(r, 2).value
        res = sh.cell(r, 3).value
        num_tuple = (num1, num2, res)
        num_list.append(num_tuple)
    return num_list

    #print(num_list[])


#function_operation()
